"""Inspect actual CSV cells and XLSX types; master data stays unchanged."""

import csv
import json
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from mailanalyst.exports.dispatch import write_output
from mailanalyst.exports.markdown import write_markdown_dataset
from mailanalyst.checks.preflight import PreflightResult, write_preflight_report
from mailanalyst.checks.system import SystemCheckResult, write_system_check_report


class SpreadsheetSafetyTests(unittest.TestCase):
    def setUp(self):
        temp = tempfile.TemporaryDirectory()
        self.addCleanup(temp.cleanup)
        self.root = Path(temp.name)
        self.texts = ["=1+1", "+SUM(1,2)", "-1+2", "@SUM(A1)", "  =1+1", "\ttext", "\r=1+1", "\n=1+1",
                      "＝1+1", "https://example.test", "Normaler Text", "#N/A"]
        self.frame = pd.DataFrame({"subject": self.texts, "number": [-2] * len(self.texts)})

    def test_csv_protects_dangerous_texts_preserving_numbers_and_master(self):
        before = self.frame.copy(deep=True)
        path = self.root / "review.csv"
        write_output(self.frame, path)
        with path.open(encoding="utf-8-sig", newline="") as file:
            rows = list(csv.DictReader(file))
        self.assertEqual([row["subject"] for row in rows], ["'" + value for value in self.texts[:9]] + self.texts[9:])
        self.assertTrue(all(row["number"] == "-2" for row in rows))
        pd.testing.assert_frame_equal(self.frame, before)
        write_output(self.frame, self.root / "master.json")
        write_output(self.frame, self.root / "master.parquet")
        self.assertEqual([row["subject"] for row in json.loads((self.root / "master.json").read_text(encoding="utf-8"))], self.texts)
        pd.testing.assert_frame_equal(pd.read_parquet(self.root / "master.parquet"), before)

    def test_excel_has_no_formulas_or_hyperlinks_and_preserves_strings(self):
        path = self.root / "review.xlsx"
        write_output(self.frame, path)
        workbook = load_workbook(path, data_only=False)
        self.addCleanup(workbook.close)
        sheet = workbook.active
        self.assertEqual([sheet.cell(index + 2, 1).value for index in range(len(self.texts))],
                         [value.replace("\r", "\n") for value in self.texts])
        for row in sheet.iter_rows():
            for cell in row:
                self.assertNotEqual(cell.data_type, "f")
                self.assertIsNone(cell.hyperlink)
        self.assertTrue(all(sheet.cell(index + 2, 1).data_type == "s" for index in range(len(self.texts))))
        self.assertEqual(sheet.cell(2, 2).data_type, "n")

    def test_markdown_csv_and_reports_are_protected_but_json_is_original(self):
        frame = pd.DataFrame([{"subject": "=1+1", "source_path": "@source", "body_preview": "+preview"}])
        write_markdown_dataset(frame, self.root / "workspace")
        with (self.root / "workspace/index.csv").open(encoding="utf-8-sig", newline="") as file:
            row = next(csv.DictReader(file))
        self.assertEqual(row["subject"], "'=1+1")
        index = json.loads((self.root / "workspace/index.jsonl").read_text(encoding="utf-8"))
        self.assertEqual(index["subject"], "=1+1")
        write_preflight_report([PreflightResult("@source", ".eml", 1, 0, "error", "=reason", False)], self.root)
        write_system_check_report([SystemCheckResult("test", "test", "warning", "=detail")], self.root)
        for filename, key in (("preflight_report.csv", "reason"), ("system_check_report.csv", "detail")):
            with (self.root / filename).open(encoding="utf-8-sig", newline="") as file:
                self.assertTrue(next(csv.DictReader(file))[key].startswith("'="))

    def test_csv_prefix_handles_leading_control_characters(self):
        from mailanalyst.text.cells import csv_text
        for value in ("\x00=1+1", "\ufeff=1+1", "\u00a0\x00  =1+1"):
            self.assertEqual(csv_text(value), "'" + value)
